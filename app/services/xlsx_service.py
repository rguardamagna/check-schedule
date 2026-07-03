import openpyxl
import re
from datetime import datetime, date, timedelta
from app import db
from app.models import Cheque


def parsear_columnas(headers):
    mapping = {}
    for i, h in enumerate(headers):
        hn = str(h).strip().lower()
        if "fecha" in hn and "pago" in hn:
            mapping["fecha"] = i
        elif hn in ("importe",) or "importe" in hn:
            mapping["importe"] = i
        elif "proveedor" in hn or "nombre" in hn:
            mapping["proveedor"] = i
        elif hn in ("estado",) or "estado" in hn:
            mapping["estado"] = i
    return mapping


def _parsear_fecha(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, (int, float)):
        return date(1899, 12, 30) + timedelta(days=int(valor))
    for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
        try:
            return datetime.strptime(str(valor).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _parsear_importe(valor):
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)

    try:
        texto = str(valor).strip()
        # Sacar TODO lo que no sea dígito, punto, coma o guión
        limpio = re.sub(r"[^\d,.\-]", "", texto)
        if not limpio:
            return None

        # Detecta formato argentino: la ULTIMA coma es decimal
        # ej: "436.666,67" -> ultima coma en pos 7, ultimo punto en pos 3
        # ej: "1250,50" -> coma pero sin punto
        # ej: "5,000.00" -> ultimo punto despues de la coma (formato US)
        if "," in limpio and limpio.rfind(",") > limpio.rfind("."):
            # Formato argentino: . separador miles, , decimal
            limpio = limpio.replace(".", "").replace(",", ".")
        else:
            # Formato internacional: , separador miles o sin separador
            limpio = limpio.replace(",", "")

        return float(limpio)
    except (ValueError, AttributeError):
        return None


def procesar_xlsx(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        wb.close()
        return {
            "importados": 0,
            "duplicados": 0,
            "errores": [{"fila": 0, "detalle": "Archivo vacío"}],
            "total_en_db": Cheque.query.count(),
        }

    headers = [str(c) if c else "" for c in rows[0]]
    mapping = parsear_columnas(headers)

    errores = []
    importados = 0
    duplicados = 0

    if "fecha" not in mapping or "importe" not in mapping or "proveedor" not in mapping:
        faltan = [k for k in ["fecha", "importe", "proveedor"] if k not in mapping]
        wb.close()
        return {
            "importados": 0,
            "duplicados": 0,
            "errores": [
                {
                    "fila": 1,
                    "detalle": f"Faltan: {', '.join(faltan)}. Headers: {headers}",
                }
            ],
            "total_en_db": Cheque.query.count(),
        }

    for idx, row in enumerate(rows[1:], start=2):
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        try:
            fecha = _parsear_fecha(row[mapping["fecha"]])
            importe = _parsear_importe(row[mapping["importe"]])
            proveedor = str(row[mapping["proveedor"]] or "").strip()

            if not fecha:
                errores.append(
                    {"fila": idx, "detalle": f"Fecha inválida: {row[mapping['fecha']]}"}
                )
                continue
            if not importe or importe < 0:
                errores.append(
                    {
                        "fila": idx,
                        "detalle": f"Importe inválido: {row[mapping['importe']]}",
                    }
                )
                continue
            if not proveedor:
                errores.append({"fila": idx, "detalle": "Proveedor vacío"})
                continue

            if Cheque.query.filter_by(
                proveedor=proveedor, importe=importe, fecha_vencimiento=fecha
            ).first():
                duplicados += 1
                continue

            estado_col = mapping.get("estado", -1)
            estado = "pendiente"
            if estado_col >= 0 and row[estado_col]:
                val = str(row[estado_col]).strip().lower()
                if val in ("pagado", "pendiente", "vencido"):
                    estado = val

            db.session.add(
                Cheque(
                    proveedor=proveedor,
                    importe=importe,
                    fecha_vencimiento=fecha,
                    estado=estado,
                    fecha_pago=fecha if estado == "pagado" else None,
                )
            )
            importados += 1

        except Exception as e:
            errores.append({"fila": idx, "detalle": f"Error: {str(e)}"})
            continue

    db.session.commit()
    wb.close()

    return {
        "importados": importados,
        "duplicados": duplicados,
        "errores": errores,
        "total_en_db": Cheque.query.count(),
    }
